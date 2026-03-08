import openpyxl

# Load the template workbook
wb = openpyxl.load_workbook('Seasonal Training Plan Workbook - NCCP - Template.xlsm')

# Get the active sheet or the sheet you want to work with
sheet = wb.active

# Fill in the Planner cells
sheet['E1'] = 'Your finalized content for E1'
# Update the other specified cells accordingly
sheet['B6'] = 'Your content for B6'
sheet['B8'] = 'Your content for B8'

# Objectives
sheet['A17'] = 'Your objectives content A17'
sheet['B17'] = 'Your objectives content B17'

# Competitions
sheet['A18'] = 'Your competitions content A18'
sheet['B18'] = 'Your competitions content B18'

# Events
sheet['D27'] = 'Your events content D27'
sheet['E27'] = 'Your events content E27'

# Techniques, Tactics, Athletic
sheet['G27'] = 'Your techniques content G27'
sheet['I27'] = 'Your tactics content I27'
sheet['K27'] = 'Your athletic content K27'

# Color coding week blocks (assuming you have cell blocks defined)
# This is a placeholder for the actual implementation based on your format
# Example of setting colors:
for row in range(1, 10):  # Assuming week block rows
    for col in range(1, 5):  # Placeholder range
        cell = sheet.cell(row=row, column=col)
        if some_condition:
            cell.fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
        elif some_other_condition:
            cell.fill = openpyxl.styles.PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')  # Blue
        else:
            cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow

# Save the workbook with the new file name
wb.save('STP_2026_U13_Intermediate_Singles_Final.xlsm')
